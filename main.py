import tkinter as tk
from tkinter import ttk
import openpyxl
import tkinter.messagebox as messagebox
from tkinter.simpledialog import askstring
def load_data():
    path = "./people.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    print(list_values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)

def validate_price_input(P):
    if P == "" or P.isdigit():
        return True
    else:
        return False
def insert_row():
    name = name_entry.get()
    age = price_entry.get()
    subscription_status = status_combobox.get()
    employment_status = "Employed" if a.get() else "Unemployed"

    if not age.isdigit():
        messagebox.showwarning("Cảnh báo", "Giá tiền phải là một số nguyên dương.")
        return

    if name == "" or age == "":
        messagebox.showwarning("Cảnh báo", "Vui lòng nhập tên sản phẩm và giá tiền.")
        return

    print(name, age, subscription_status, employment_status)

    # Insert row into Excel sheet
    path = "./people.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [name, age, subscription_status, employment_status]
    sheet.append(row_values)
    
    workbook.save(path)

    # Insert row into treeview
    treeview.insert('', tk.END, values=row_values)
    
    # Clear the values
    name_entry.delete(0, "end")
    name_entry.insert(0, "Nhập tên sản phẩm")
    price_entry.delete(0, "end")
    price_entry.insert(0, "Nhập giá tiền")
    status_combobox.set(combo_list[0])
    checkbutton.state(["!selected"])




# def toggle_mode():
#     if mode_switch.instate(["selected"]):
#         style.theme_use("forest-light")
#     else:
#         style.theme_use("forest-dark")


root = tk.Tk()



root.geometry("1380x740")
style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-light")

combo_list = ["Thức ăn chăn nuôi", "Rau", "Hoa quả"]

frame = ttk.Frame(root)
frame.pack()

style = ttk.Style()
style.configure("TEntry.CustomHeight.TEntry", padding=[10, 10], fieldbackground="white", height=40,)  # Tùy chỉnh chiều cao
style.configure("TCombobox.CustomHeight.TCombobox", padding=[10, 10], fieldbackground="white", height=40)

widgets_frame = ttk.LabelFrame(frame, text="Thônng tin sản phẩm")
widgets_frame.grid(row=0, column=0, padx=20, pady=10)


name_entry = ttk.Entry(widgets_frame,style="TEntry.CustomHeight.TEntry",width=80)
name_entry.insert(0, "Nhập tên sản phẩm")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
name_entry.grid(row=0, column=0, padx=5, pady=(30, 5), sticky="ew")

price_entry = ttk.Entry(widgets_frame,style="TEntry.CustomHeight.TEntry")
price_entry.insert(0, "Nhập giá tiền")
price_entry.bind("<FocusIn>", lambda e: price_entry.delete('0', 'end'))
price_entry.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

status_combobox = ttk.Combobox(widgets_frame, values=combo_list,style="TCombobox.CustomHeight.TCombobox")
status_combobox.current(0)
status_combobox.grid(row=2, column=0, padx=5, pady=5,  sticky="ew")

a = tk.BooleanVar()
checkbutton = ttk.Checkbutton(widgets_frame, text="Employed", variable=a)
checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

button = ttk.Button(widgets_frame, text="Insert", command=insert_row)
button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

separator = ttk.Separator(widgets_frame)
separator.grid(row=5, column=0, padx=(20, 10), pady=10, sticky="ew")

# mode_switch = ttk.Checkbutton(
#     widgets_frame, text="Mode", style="Switch", command=toggle_mode)
# mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10,padx=10,rowspan=2)

treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Sản phẩm", "Giá tiền", "Loại", "Số lượng", "Tổng tiền")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=cols, height=13,)

treeview.column("Sản phẩm", width=130,anchor="center")  # Đặt chiều rộng cho cột "Sản phẩm" là 100
treeview.column("Giá tiền", width=130,anchor="center")    # Đặt chiều rộng cho cột "Giá tiền" là 80
treeview.column("Loại", width=130,anchor="center")       # Đặt chiều rộng cho cột "Loại" là 120
treeview.column("Số lượng", width=130,anchor="center") 
treeview.column("Tổng tiền", width=130,anchor="center") 

treeview.pack()
treeScroll.config(command=treeview.yview)
load_data()

# xoa dong trong treeview   
# Hàm để xóa dòng được chọn từ treeview
def delete_selected_row():
    selected_item = treeview.selection()
    for item in selected_item:
        treeview.delete(item)
        delete_row_from_excel(item)

# Hàm để xóa dòng khỏi tệp Excel
def delete_row_from_excel(selected_item):
    path = "./people.xlsx"
    
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    for item in selected_item:
        row_index = int(treeview.index(item))  # Lấy chỉ số hàng
        sheet.delete_rows(row_index)
    workbook.save(path)





def on_treeview_select(event):
    delete_button.config(state=tk.NORMAL if treeview.selection() else tk.DISABLED)
# new
buttonFrame = ttk.Frame(frame)
buttonFrame.grid(row=1, column=0, pady=10,padx=10)
treeview.bind("<<TreeviewSelect>>", on_treeview_select)
delete_button = tk.Button(buttonFrame, text="Xoá", command=delete_selected_row, state=tk.DISABLED)
delete_button.pack()
# update button
def update_selected_row():
    selected_item = treeview.selection()
    if selected_item:
        current_data = treeview.item(selected_item, 'values')
        new_data = askstring("Chỉnh sửa thông tin", "Nhập thông tin mới (Name, Age):", initialvalue="{}, {}".format(current_data[0], current_data[1]))
        if new_data:
            new_data = new_data.split(',')
            if len(new_data) == 2:
                treeview.item(selected_item, values=(new_data[0].strip(), new_data[1].strip()))
    

def on_treeview_select(event):
    update_button.config(state=tk.NORMAL if treeview.selection() else tk.DISABLED)

update_button = tk.Button(buttonFrame, text="Cập nhật", command=update_selected_row)

update_button.pack()
root.mainloop()
